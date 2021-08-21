"""
Converter module
"""

import csv
import uuid
import os.path
from datetime import date

import openpyxl
import pandas as pd

from common.app import App
from common.database import Database
from common.helpers import Helpers


class Converter:
    """
    Convert database table to csv
    """

    def __init__(self, app: App, database: Database, only_covid: bool):
        self.app = app
        self.database = database
        self.only_covid = only_covid

    def __file_exists(self, file_name):
        """
        Check if a file of that name already exists in the database/xlsx folder
        """

        return bool(
            os.path.isfile(f"{self.app.root_dir}/database/xlsx/{file_name}.xlsx")
        )

    def __dir_exists(self, dirname: str) -> bool:
        """
        Check if directory exists in database folder
        """

        return bool(os.path.isdir(f"{self.app.root_dir}/database/{dirname}"))

    def __gen_csv_reader(self, file):
        with open(
            f"{self.app.root_dir}/database/csv/{file}", encoding="utf8", newline=""
        ) as f:
            reader = csv.reader(f)
            for row in reader:
                yield row

    def convert_by_columns(self, cols=tuple(col for col in Helpers.schema_cols)):
        """
        Convert tweets table to csv by columns
        """

        # Check if database/csv/ dir exists
        if not self.__dir_exists("csv"):
            os.mkdir(f"{self.app.root_dir}/database/csv/")

        today = str(date.today())

        covid = "Covid-" if self.only_covid else "Tot-"
        outfile = self.database.db_name
        outfile = outfile.strip(".db")

        with self.database, open(
            f"{self.app.root_dir}/database/csv/{covid}{outfile}-{today}.csv",
            "w+",
            encoding="utf-8",
            newline="",
        ) as out:
            try:
                writer = csv.writer(out, delimiter=",", quoting=csv.QUOTE_NONNUMERIC)
                writer.writerow(cols)
                tweets = self.database.get_fields(list(cols), self.only_covid)
                writer.writerows(tweets)
                print(
                    f"Table successfully converted as database/csv/{covid}{outfile}-{today}.csv"
                )

                return f"{covid}{outfile}-{today}.csv"
            except csv.Error as error:
                print("Error while writing CSV", error)

    def csv_to_xlsx(self, csv_file):
        """
        Export a csv file to xlsx
        """

        # Check if database/xlsx/ dir exists
        if not self.__dir_exists("xlsx"):
            os.mkdir(f"{self.app.root_dir}/database/xlsx/")

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        with open(
            f"{self.app.root_dir}/database/csv/{csv_file}",
            "r",
            encoding="utf8",
            newline="",
        ) as file:
            reader = csv.reader(file)
            for row_i, row in enumerate(reader, start=1):
                for col_i, val in enumerate(row, start=1):
                    sheet.cell(row=row_i, column=col_i).value = val
        try:
            csv_file = csv_file.strip(".csv")
            orig_name = csv_file
            if self.__file_exists(csv_file):
                csv_file = orig_name + "-" + str(uuid.uuid4())[:7]
                if self.__file_exists(csv_file):
                    csv_file = orig_name + "-" + str(uuid.uuid4())[:7]

            workbook.save(f"{self.app.root_dir}/database/xlsx/{csv_file}.xlsx")
            print(f"File successfully exported to database/xlsx/{csv_file}.xlsx")

            return f"{csv_file}.xlsx"
        except openpyxl.utils.exceptions.InvalidFileException as error:
            print("Error", error)

    def csv_to_xlsx_v2(self, csv_file):
        """
        Export a csv file to xlsx
        New implementation, less memory-hungry.
        """

        # Check if database/xlsx/ dir exists
        if not self.__dir_exists("xlsx"):
            os.mkdir(f"{self.app.root_dir}/database/xlsx/")

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        with open(
            f"{self.app.root_dir}/database/csv/{csv_file}",
            "r",
            encoding="utf8",
            newline="",
        ) as file:
            reader = csv.reader(file)
            for row in reader:
                sheet.append(row)
        try:
            csv_file = csv_file.strip(".csv")
            orig_name = csv_file

            if self.__file_exists(csv_file):
                csv_file = orig_name + "-" + str(uuid.uuid4())[:7]
                if self.__file_exists(csv_file):
                    csv_file = orig_name + "-" + str(uuid.uuid4())[:7]

            workbook.save(f"{self.app.root_dir}/database/xlsx/{csv_file}.xlsx")
            print(f"File successfully exported to database/xlsx/{csv_file}.xlsx")

            return f"{csv_file}.xlsx"
        except openpyxl.utils.exceptions.InvalidFileException as error:
            print("Error", error)

    def csv_to_xlsx_v3(self, csv_file):
        """
        Export a csv file to xlsx
        New implementation, less memory-hungry.
        """

        # Check if database/xlsx/ dir exists
        if not self.__dir_exists("xlsx"):
            os.mkdir(f"{self.app.root_dir}/database/xlsx/")

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        csv_gen = self.__gen_csv_reader(csv_file)
        for row in csv_gen:
            # print(row)
            sheet.append(row)

        try:
            csv_file = csv_file.strip(".csv")
            orig_name = csv_file

            if self.__file_exists(csv_file):
                csv_file = orig_name + "-" + str(uuid.uuid4())[:7]
                if self.__file_exists(csv_file):
                    csv_file = orig_name + "-" + str(uuid.uuid4())[:7]

            workbook.save(f"{self.app.root_dir}/database/xlsx/{csv_file}.xlsx")
            print(f"File successfully exported to database/xlsx/{csv_file}.xlsx")

            return f"{csv_file}.xlsx"
        except openpyxl.utils.exceptions.InvalidFileException as error:
            print("Error", error)

    def csv_to_xlsx_pd(self, csv_file):
        """
        Export a csv file to xlsx
        New implementation, less memory-hungry.
        """

        # Check if database/xlsx/ dir exists
        if not self.__dir_exists("xlsx"):
            os.mkdir(f"{self.app.root_dir}/database/xlsx/")

        df = pd.read_csv(f"{self.app.root_dir}/database/csv/{csv_file}")

        csv_file = csv_file.strip(".csv")
        orig_name = csv_file

        if self.__file_exists(csv_file):
            csv_file = orig_name + "-" + str(uuid.uuid4())[:7]
            if self.__file_exists(csv_file):
                csv_file = orig_name + "-" + str(uuid.uuid4())[:7]

        df.to_excel(f"{self.app.root_dir}/database/xlsx/{csv_file}.xlsx", index=False)
        print(f"File successfully exported to database/xlsx/{csv_file}.xlsx")

        return f"{csv_file}.xlsx"

    def csv_to_xlsx_pd_v2(self, csv_file):
        """
        Export a csv file to xlsx
        New implementation, less memory-hungry.
        """

        # Check if database/xlsx/ dir exists
        if not self.__dir_exists("xlsx"):
            os.mkdir(f"{self.app.root_dir}/database/xlsx/")

        df = pd.read_csv(f"{self.app.root_dir}/database/csv/{csv_file}")

        csv_file = csv_file.strip(".csv")
        orig_name = csv_file

        if self.__file_exists(csv_file):
            csv_file = orig_name + "-" + str(uuid.uuid4())[:7]
            if self.__file_exists(csv_file):
                csv_file = orig_name + "-" + str(uuid.uuid4())[:7]

        writer = pd.ExcelWriter(
            f"{self.app.root_dir}/database/xlsx/{csv_file}.xlsx", enginge="xlswriter"
        )
        df.to_excel(writer, index=False)
        writer.save()

        print(f"File successfully exported to database/xlsx/{csv_file}.xlsx")

        return f"{csv_file}.xlsx"

    def csv_to_xlsx_pyexcel(self, csv_file):
        """
        Export a csv file to xlsx
        New implementation, less memory-hungry.
        """

        import pyexcel

        # Check if database/xlsx/ dir exists
        if not self.__dir_exists("xlsx"):
            os.mkdir(f"{self.app.root_dir}/database/xlsx/")

        sheet = pyexcel.get_sheet(
            file_name=f"{self.app.root_dir}/database/csv/{csv_file}", delimiter=","
        )

        csv_file = csv_file.strip(".csv")
        orig_name = csv_file

        if self.__file_exists(csv_file):
            csv_file = orig_name + "-" + str(uuid.uuid4())[:7]
            if self.__file_exists(csv_file):
                csv_file = orig_name + "-" + str(uuid.uuid4())[:7]

        sheet.save_as(f"{self.app.root_dir}/database/xlsx/{csv_file}.xlsx")
        print(f"File successfully exported to database/xlsx/{csv_file}.xlsx")

        return f"{csv_file}.xlsx"
